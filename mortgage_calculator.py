import pandas as pd
import numpy as np
from math import ceil
from datetime import datetime, date
import calendar
from datetime import date, timedelta
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, numbers
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import string
from tkinter import *
from tkinter import ttk
from tkinter import messagebox

##################################################
def roundup2d(number):
    rounded = ceil(100*number)/100
    return rounded

def calculate_monthly_payment(length_in_year, annual_rate, total_amount):
    monthly_rate = annual_rate/12
    length_in_month = length_in_year*12
    monthly_payment = (total_amount*monthly_rate)/(1-1/(1+monthly_rate)**length_in_month)
    monthly_payment = roundup2d(monthly_payment)
    return monthly_payment

# Explanation of monthly payment
# If annual rate is 6% and payment is made 10-Year long, then the monthly rate is 0.5% and it rolls over 120 periods by (1.005)^120.

def payment_table(length_in_year, annual_rate, total_amount, start_date=datetime.now()):
    payment_table = pd.DataFrame(columns=['Payment Date','Original Balance','Payment','Interest','Balance Deduction','New Balance'])
    monthly_rate = annual_rate/12
    monthly_payment = calculate_monthly_payment(length_in_year,annual_rate,total_amount)
    payment = monthly_payment
    original_balance = total_amount
    payment_date = start_date
    
    for x in list(range(length_in_year*12)):
        interest = round(original_balance*monthly_rate,2)
        # interest = roundup2d(original_balance*monthly_rate)
        payment = min(monthly_payment, original_balance+interest)
        balance_deduction = payment-interest
        new_balance = original_balance - balance_deduction
        payment_date = payment_date + timedelta(days=calendar.monthrange(payment_date.year, payment_date.month)[1])
        payment_table.loc[x] = (payment_date.strftime("%Y-%m-%d"), -1*original_balance, payment, interest, balance_deduction, -1*new_balance)
        original_balance = new_balance
    return payment_table

def center(cell):
    cell.alignment = Alignment(horizontal='center')
    return None

########################################
def start_calculation():
    length_in_year = input1_entry.get()
    annual_rate = input2_entry.get()
    total_amount = input3_entry.get()
    loan_date = loan_date_entry.get()
    try:
        length_in_year = int(length_in_year)
        annual_rate = float(annual_rate)/100
        total_amount = float(total_amount)
        monthly_payment = calculate_monthly_payment(length_in_year,annual_rate,total_amount)
        table_1 = payment_table(length_in_year, annual_rate, total_amount, start_date=date.fromisoformat(loan_date))
        total_payment = round(np.sum(table_1['Payment']),2)
        total_interest = round(np.sum(table_1['Interest']),2)
        interest_ratio = round(total_interest/total_payment,4)
        label_61.config(text = f"{monthly_payment:<0.2f}")
        label_71.config(text = f"{total_payment:<0.2f}")
        label_81.config(text = f"{total_interest:<0.2f}")
        label_91.config(text = f"{100*interest_ratio:<0.2f}%")
    except ValueError:
        label_61.config(text = "Input Value Error")
        label_71.config(text = "Input Value Error")
        label_81.config(text = "Input Value Error")
        label_91.config(text = "Input Value Error")
    return None

def export_schedule():
    length_in_year = input1_entry.get()
    annual_rate = input2_entry.get()
    total_amount = input3_entry.get()
    loan_date = loan_date_entry.get()
    label_11_1.config(text = "Export in progress")
    try:
        length_in_year = int(length_in_year)
        annual_rate = float(annual_rate)/100
        total_amount = float(total_amount)
        monthly_payment = calculate_monthly_payment(length_in_year,annual_rate,total_amount)
        table_1 = payment_table(length_in_year, annual_rate, total_amount, start_date=date.fromisoformat(loan_date))
        total_payment = round(np.sum(table_1['Payment']),2)
        total_interest = round(np.sum(table_1['Interest']),2)
        interest_ratio = round(total_interest/total_payment,4)
        export_url = r'C:\Users\User\Desktop\Mortgage_Result.xlsx'
        table_1.to_excel(export_url, index = False, header = True)
        wb = openpyxl.load_workbook(export_url)
        ws = wb["Sheet1"]
    
        # Add information
        ws["H1"], ws["I1"], ws["J1"], ws["K1"] = "Loan Amount", "Annual Rate", "Num of years", "Loan Date"
        ws["H2"], ws["I2"], ws["J2"], ws["K2"] = total_amount, annual_rate, length_in_year, loan_date
        ws["I2"].number_format = '0.00%'
        ws["H4"], ws["I4"], ws["J4"], ws["K4"] = "Monthly Payment", "Total Payment", "Total Interest", "Interest Ratio"
        ws["H5"], ws["I5"], ws["J5"], ws["K5"] = monthly_payment, total_payment, total_interest, interest_ratio
        ws["K5"].number_format = '0.00%'
    
        # Add thousand separator and center alignment
        for cell in [ws["H2"], ws["H5"], ws["I5"], ws["J5"]]:
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        for i in np.array(range(ws.max_row))+1:
            for j in np.array(range(5))+2:
                ws.cell(row=i, column=j).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            for j in np.array(range(11))+1:
                center(ws.cell(row=i, column=j))
        # Change column width
        for i in list(string.ascii_uppercase)[0:11]:
            ws.column_dimensions[i].width = 20
        
        wb.save(export_url)
        label_11_1.config(text = "Exported to Desktop")
    except ValueError:
        label_61.config(text = "Input Value Error")
        label_71.config(text = "Input Value Error")
        label_81.config(text = "Input Value Error")
        label_91.config(text = "Input Value Error")
        label_11_1.config(text = "Export Error")
    except NameError:
        label_11_1.config(text = "Export Error")
    except PermissionError:
        label_11_1.config(text = "Export Error")
        messagebox.showerror("Export Failed", f"File <{export_url}> already opened. Please close it before running.")
    return None

##############################
root = Tk()
root.title('Mortgage Calculator')
frm = ttk.Frame(root, padding=100)
frm.grid()
label_title = ttk.Label(frm, text="Mortgage Calculator")
label_title.grid(row=0, column=1)

label_10 = ttk.Label(frm, text="")
label_10.grid(row=1, column=0)

label_input1 = ttk.Label(frm, text="Length in years:")
label_input1.grid(row=2, column=0)
label_input2 = ttk.Label(frm, text="Annual int rate (%):")
label_input2.grid(row=3, column=0)
label_input3 = ttk.Label(frm, text="Loan amount:")
label_input3.grid(row=4, column=0)
input1_entry = ttk.Entry(frm)
input1_entry.grid(row=2, column=1)
input2_entry = ttk.Entry(frm)
input2_entry.grid(row=3, column=1)
input3_entry = ttk.Entry(frm)
input3_entry.grid(row=4, column=1)
label_loan_date = ttk.Label(frm, text="Loan Date:")
label_loan_date.grid(row=5, column=0)
loan_date_entry = ttk.Entry(frm)
loan_date_entry.grid(row=5, column=1)
loan_date_entry.insert(0,datetime.now().strftime("%Y-%m-%d"))


label_60 = ttk.Label(frm, text="Monthly Payment:")
label_60.grid(row=6, column=0)
label_61 = ttk.Label(frm, text="")
label_61.grid(row=6, column=1)
label_70 = ttk.Label(frm, text="Total Payment:")
label_70.grid(row=7, column=0)
label_71 = ttk.Label(frm, text="")
label_71.grid(row=7, column=1)
label_80 = ttk.Label(frm, text="Total Interest:")
label_80.grid(row=8, column=0)
label_81 = ttk.Label(frm, text="")
label_81.grid(row=8, column=1)
label_90 = ttk.Label(frm, text="Interest Ratio:")
label_90.grid(row=9, column=0)
label_91 = ttk.Label(frm, text="")
label_91.grid(row=9, column=1)

ttk.Button(frm, text="Calculate", command=start_calculation).grid(column=1, row=10)

label_11_0 = ttk.Label(frm, text="")
label_11_0.grid(row=11, column=0)
label_11_1 = ttk.Label(frm, text="")
label_11_1.grid(row=11, column=1)


ttk.Button(frm, text="Export Schedule", command=export_schedule).grid(column=1, row=12)

root.mainloop()